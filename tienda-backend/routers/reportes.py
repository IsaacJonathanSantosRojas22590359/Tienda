from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from database import get_connection
from auth import solo_admin
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter(prefix='/api', tags=['Reportes'])

@router.get('/reportes/dashboard/')
def dashboard(payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT COALESCE(SUM(total), 0) AS ingresos_mes
        FROM ventas
        WHERE MONTH(fecha) = MONTH(NOW()) AND YEAR(fecha) = YEAR(NOW())
    """)
    ingresos_mes = float(cursor.fetchone()['ingresos_mes'])

    cursor.execute("""
        SELECT COUNT(*) AS ventas_mes FROM ventas
        WHERE MONTH(fecha) = MONTH(NOW()) AND YEAR(fecha) = YEAR(NOW())
    """)
    ventas_mes = cursor.fetchone()['ventas_mes']

    cursor.execute("SELECT COALESCE(SUM(stock), 0) AS total FROM productos WHERE activo = TRUE")
    total_stock = int(cursor.fetchone()['total'])

    cursor.execute("SELECT COUNT(*) AS bajo FROM productos WHERE stock < 10 AND activo = TRUE")
    stock_bajo = cursor.fetchone()['bajo']

    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE activo = TRUE")
    empleados = cursor.fetchone()['total']

    cursor.execute("""
        SELECT v.id, u.nombre AS empleado, v.total, v.fecha, v.metodo_pago
        FROM ventas v
        INNER JOIN usuarios u ON v.usuario_id = u.id
        ORDER BY v.fecha DESC LIMIT 5
    """)
    ultimas = cursor.fetchall()
    for v in ultimas:
        v['total'] = float(v['total'])
        v['fecha'] = str(v['fecha'])

    cursor.execute("""
        SELECT p.nombre, SUM(dv.cantidad) AS unidades, SUM(dv.subtotal) AS ingresos
        FROM detalle_ventas dv
        INNER JOIN productos p ON dv.producto_id = p.id
        GROUP BY p.id, p.nombre
        ORDER BY unidades DESC LIMIT 5
    """)
    top = cursor.fetchall()
    for t in top:
        t['unidades'] = int(t['unidades'])
        t['ingresos'] = float(t['ingresos'])

    cursor.close(); conn.close()
    return {
        'ingresos_mes':   ingresos_mes,
        'ventas_mes':     ventas_mes,
        'total_stock':    total_stock,
        'stock_bajo':     stock_bajo,
        'empleados':      empleados,
        'ultimas_ventas': ultimas,
        'top_productos':  top,
    }

@router.get('/reportes/ventas-por-dia/')
def ventas_por_dia(desde: str = Query(...), hasta: str = Query(...),
                   payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT DATE(fecha) AS dia, COUNT(*) AS total_ventas, SUM(total) AS ingresos
        FROM ventas
        WHERE DATE(fecha) BETWEEN %s AND %s
        GROUP BY DATE(fecha) ORDER BY dia ASC
    """, (desde, hasta))
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    for r in rows:
        r['dia']      = str(r['dia'])
        r['ingresos'] = float(r['ingresos'])
    return rows

@router.get('/reportes/ventas-por-mes/')
def ventas_por_mes(payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT MONTH(fecha) AS mes, COUNT(*) AS total_ventas, SUM(total) AS ingresos
        FROM ventas WHERE YEAR(fecha) = YEAR(NOW())
        GROUP BY MONTH(fecha) ORDER BY mes ASC
    """)
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    meses = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    return [{'mes': meses[r['mes']-1], 'total_ventas': r['total_ventas'],
             'ingresos': float(r['ingresos'])} for r in rows]

@router.get('/reportes/top-productos/')
def top_productos(limite: int = 10, payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT p.nombre, c.nombre AS categoria,
               SUM(dv.cantidad) AS unidades, SUM(dv.subtotal) AS ingresos
        FROM detalle_ventas dv
        INNER JOIN productos p ON dv.producto_id = p.id
        LEFT JOIN categorias c ON p.categoria_id = c.id
        GROUP BY p.id, p.nombre, c.nombre
        ORDER BY unidades DESC LIMIT %s
    """, (limite,))
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    for r in rows:
        r['unidades'] = int(r['unidades'])
        r['ingresos'] = float(r['ingresos'])
    return rows

@router.get('/reportes/exportar/pdf/')
def exportar_pdf(desde: str = Query(...), hasta: str = Query(...),
                 payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT v.id, u.nombre, v.total, v.fecha, v.metodo_pago
        FROM ventas v INNER JOIN usuarios u ON v.usuario_id = u.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s ORDER BY v.fecha DESC
    """, (desde, hasta))
    ventas = cursor.fetchall()
    cursor.execute("""
        SELECT p.nombre, c.nombre AS categoria,
               SUM(dv.cantidad) AS unidades, SUM(dv.subtotal) AS ingresos
        FROM detalle_ventas dv
        INNER JOIN productos p ON dv.producto_id = p.id
        LEFT JOIN categorias c ON p.categoria_id = c.id
        INNER JOIN ventas v ON dv.venta_id = v.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        GROUP BY p.id, p.nombre, c.nombre ORDER BY unidades DESC LIMIT 10
    """, (desde, hasta))
    top = cursor.fetchall()
    cursor.execute("""
        SELECT COALESCE(SUM(total),0) AS total, COUNT(*) AS count
        FROM ventas WHERE DATE(fecha) BETWEEN %s AND %s
    """, (desde, hasta))
    resumen = cursor.fetchone()
    cursor.close(); conn.close()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter,
                               topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story  = []

    story.append(Paragraph(f'<b>Reporte de Ventas</b><br/><font size=11>Período: {desde} al {hasta}</font>', styles['Title']))
    story.append(Spacer(1, 0.25*inch))

    total_i = float(resumen['total'])
    count_v = resumen['count']
    promedio = total_i / count_v if count_v > 0 else 0
    res_data = [
        ['Métrica', 'Valor'],
        ['Total ingresos', f'${total_i:,.2f}'],
        ['Total ventas',   str(count_v)],
        ['Promedio/venta', f'${promedio:,.2f}'],
    ]
    t_res = Table(res_data, colWidths=[3*inch, 2*inch])
    t_res.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0),  colors.HexColor('#1e2a3a')),
        ('TEXTCOLOR',      (0,0), (-1,0),  colors.white),
        ('FONTNAME',       (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,-1), 10),
        ('ALIGN',          (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f0f4ff'), colors.white]),
        ('GRID',           (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',     (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 6),
    ]))
    story.append(t_res)
    story.append(Spacer(1, 0.3*inch))

    story.append(Paragraph('<b>Productos más vendidos</b>', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    top_data = [['Producto', 'Categoría', 'Unidades', 'Ingresos']]
    for r in top:
        top_data.append([r['nombre'], r['categoria'] or '—', str(int(r['unidades'])), f"${float(r['ingresos']):,.2f}"])
    t_top = Table(top_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.5*inch])
    t_top.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0),  colors.HexColor('#7b9cff')),
        ('TEXTCOLOR',      (0,0), (-1,0),  colors.white),
        ('FONTNAME',       (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,-1), 9),
        ('ALIGN',          (2,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8faff'), colors.white]),
        ('GRID',           (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',     (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 5),
    ]))
    story.append(t_top)
    story.append(Spacer(1, 0.3*inch))

    story.append(Paragraph('<b>Detalle de ventas</b>', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    v_data = [['#', 'Empleado', 'Total', 'Fecha', 'Método']]
    for v in ventas:
        v_data.append([str(v['id']), v['nombre'], f"${float(v['total']):,.2f}", str(v['fecha'])[:16], v['metodo_pago']])
    t_v = Table(v_data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 1.8*inch, 1*inch])
    t_v.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0),  colors.HexColor('#1e2a3a')),
        ('TEXTCOLOR',      (0,0), (-1,0),  colors.white),
        ('FONTNAME',       (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,-1), 8),
        ('ALIGN',          (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f0f4ff'), colors.white]),
        ('GRID',           (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',     (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 4),
    ]))
    story.append(t_v)
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f'<font size=8>Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}</font>', styles['Normal']))
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="reporte_{desde}_{hasta}.pdf"'})

@router.get('/reportes/exportar/excel/')
def exportar_excel(desde: str = Query(...), hasta: str = Query(...),
                   payload: dict = Depends(solo_admin)):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT v.id, u.nombre, v.total, v.fecha, v.metodo_pago
        FROM ventas v INNER JOIN usuarios u ON v.usuario_id = u.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s ORDER BY v.fecha DESC
    """, (desde, hasta))
    ventas = cursor.fetchall()
    cursor.execute("""
        SELECT p.nombre, c.nombre AS categoria,
               SUM(dv.cantidad) AS unidades, SUM(dv.subtotal) AS ingresos
        FROM detalle_ventas dv
        INNER JOIN productos p ON dv.producto_id = p.id
        LEFT JOIN categorias c ON p.categoria_id = c.id
        INNER JOIN ventas v ON dv.venta_id = v.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        GROUP BY p.id, p.nombre, c.nombre ORDER BY unidades DESC
    """, (desde, hasta))
    top = cursor.fetchall()
    cursor.close(); conn.close()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Ventas'
    h_fill = PatternFill('solid', fgColor='1e2a3a')
    h_font = Font(color='FFFFFF', bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center')
    alt    = PatternFill('solid', fgColor='f0f4ff')

    ws1.append(['Reporte de Ventas', '', '', '', ''])
    ws1.merge_cells('A1:E1')
    ws1['A1'].font = Font(bold=True, size=14, color='1e2a3a')
    ws1['A1'].alignment = center
    ws1.append([f'Período: {desde} al {hasta}', '', '', '', ''])
    ws1.merge_cells('A2:E2')
    ws1['A2'].alignment = center
    ws1.append([])

    enc = ['#', 'Empleado', 'Total ($)', 'Fecha', 'Método']
    ws1.append(enc)
    for col in range(1, 6):
        c = ws1.cell(row=4, column=col)
        c.fill = h_fill; c.font = h_font; c.alignment = center

    for i, v in enumerate(ventas):
        ws1.append([v['id'], v['nombre'], float(v['total']), str(v['fecha'])[:16], v['metodo_pago']])
        if i % 2 == 0:
            for col in range(1, 6):
                ws1.cell(row=5+i, column=col).fill = alt

    for col, w in zip('ABCDE', [8, 25, 15, 20, 18]):
        ws1.column_dimensions[col].width = w

    ws2 = wb.create_sheet('Top Productos')
    h_fill2 = PatternFill('solid', fgColor='7b9cff')
    ws2.append(['Productos más vendidos', '', '', ''])
    ws2.merge_cells('A1:D1')
    ws2['A1'].font = Font(bold=True, size=14, color='1e2a3a')
    ws2['A1'].alignment = center
    ws2.append([])
    enc2 = ['Producto', 'Categoría', 'Unidades', 'Ingresos ($)']
    ws2.append(enc2)
    for col in range(1, 5):
        c = ws2.cell(row=3, column=col)
        c.fill = h_fill2; c.font = h_font; c.alignment = center
    for i, p in enumerate(top):
        ws2.append([p['nombre'], p['categoria'] or '—', int(p['unidades']), float(p['ingresos'])])
        if i % 2 == 0:
            for col in range(1, 5):
                ws2.cell(row=4+i, column=col).fill = alt
    for col, w in zip('ABCD', [30, 20, 20, 18]):
        ws2.column_dimensions[col].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="reporte_{desde}_{hasta}.xlsx"'})
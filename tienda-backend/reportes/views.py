from django.shortcuts import render

# Create your views here.
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import connection

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    with connection.cursor() as cursor:

        # Query: total ingresos del mes actual
        cursor.execute("""
            SELECT COALESCE(SUM(total), 0)
            FROM ventas
            WHERE MONTH(fecha) = MONTH(NOW())
              AND YEAR(fecha)  = YEAR(NOW())
        """)
        ingresos_mes = float(cursor.fetchone()[0])

        # Query: total ventas del mes
        cursor.execute("""
            SELECT COUNT(*) FROM ventas
            WHERE MONTH(fecha) = MONTH(NOW())
              AND YEAR(fecha)  = YEAR(NOW())
        """)
        ventas_mes = cursor.fetchone()[0]

        # Query: total productos activos en stock
        cursor.execute("""
            SELECT COALESCE(SUM(stock), 0)
            FROM productos WHERE activo = TRUE
        """)
        total_stock = int(cursor.fetchone()[0])

        # Query: productos con stock bajo (menos de 10)
        cursor.execute("""
            SELECT COUNT(*) FROM productos
            WHERE stock < 10 AND activo = TRUE
        """)
        stock_bajo = cursor.fetchone()[0]

        # Query: total empleados activos
        cursor.execute("""
            SELECT COUNT(*) FROM usuarios WHERE activo = TRUE
        """)
        empleados = cursor.fetchone()[0]

        # Query: últimas 5 ventas
        cursor.execute("""
            SELECT v.id, u.nombre, v.total, v.fecha, v.metodo_pago
            FROM ventas v
            INNER JOIN usuarios u ON v.usuario_id = u.id
            ORDER BY v.fecha DESC
            LIMIT 5
        """)
        ultimas = [
            {
                'id':         row[0],
                'empleado':   row[1],
                'total':      float(row[2]),
                'fecha':      row[3],
                'metodo_pago':row[4],
            }
            for row in cursor.fetchall()
        ]

        # Query: top 5 productos más vendidos
        cursor.execute("""
            SELECT p.nombre,
                   SUM(dv.cantidad)              AS unidades,
                   SUM(dv.subtotal)              AS ingresos
            FROM detalle_ventas dv
            INNER JOIN productos p ON dv.producto_id = p.id
            GROUP BY p.id, p.nombre
            ORDER BY unidades DESC
            LIMIT 5
        """)
        top_productos = [
            {
                'nombre':   row[0],
                'unidades': int(row[1]),
                'ingresos': float(row[2]),
            }
            for row in cursor.fetchall()
        ]

    return Response({
        'ingresos_mes':  ingresos_mes,
        'ventas_mes':    ventas_mes,
        'total_stock':   total_stock,
        'stock_bajo':    stock_bajo,
        'empleados':     empleados,
        'ultimas_ventas':ultimas,
        'top_productos': top_productos,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ventas_por_dia(request):
    desde = request.query_params.get('desde', '')
    hasta = request.query_params.get('hasta', '')

    # Query: ventas agrupadas por día en un rango de fechas
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT DATE(fecha)       AS dia,
                   COUNT(*)          AS total_ventas,
                   SUM(total)        AS ingresos
            FROM ventas
            WHERE DATE(fecha) BETWEEN %s AND %s
            GROUP BY DATE(fecha)
            ORDER BY dia ASC
        """, [desde, hasta])
        rows = cursor.fetchall()

    data = [
        {'dia': str(row[0]), 'total_ventas': row[1], 'ingresos': float(row[2])}
        for row in rows
    ]
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ventas_por_mes(request):
    # Query: ventas agrupadas por mes del año actual
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT MONTH(fecha)      AS mes,
                   COUNT(*)          AS total_ventas,
                   SUM(total)        AS ingresos
            FROM ventas
            WHERE YEAR(fecha) = YEAR(NOW())
            GROUP BY MONTH(fecha)
            ORDER BY mes ASC
        """)
        rows = cursor.fetchall()

    meses = ['Ene','Feb','Mar','Abr','May','Jun',
             'Jul','Ago','Sep','Oct','Nov','Dic']
    data  = [
        {
            'mes':          meses[row[0] - 1],
            'total_ventas': row[1],
            'ingresos':     float(row[2])
        }
        for row in rows
    ]
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def top_productos(request):
    limite = request.query_params.get('limite', 10)

    # Query: productos más vendidos con ingresos generados
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.nombre,
                   c.nombre                      AS categoria,
                   SUM(dv.cantidad)              AS unidades,
                   SUM(dv.subtotal)              AS ingresos
            FROM detalle_ventas dv
            INNER JOIN productos p ON dv.producto_id = p.id
            LEFT JOIN  categorias c ON p.categoria_id = c.id
            GROUP BY p.id, p.nombre, c.nombre
            ORDER BY unidades DESC
            LIMIT %s
        """, [int(limite)])
        columnas = [col[0] for col in cursor.description]
        rows     = cursor.fetchall()

    data = [dict(zip(columnas, row)) for row in rows]
    for item in data:
        item['ingresos'] = float(item['ingresos'])
        item['unidades'] = int(item['unidades'])
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_pdf(request):
    desde = request.query_params.get('desde', '')
    hasta = request.query_params.get('hasta', '')

    # Obtener datos
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT v.id, u.nombre, v.total, v.fecha, v.metodo_pago
            FROM ventas v
            INNER JOIN usuarios u ON v.usuario_id = u.id
            WHERE DATE(v.fecha) BETWEEN %s AND %s
            ORDER BY v.fecha DESC
        """, [desde, hasta])
        ventas = cursor.fetchall()

        cursor.execute("""
            SELECT p.nombre, c.nombre, SUM(dv.cantidad), SUM(dv.subtotal)
            FROM detalle_ventas dv
            INNER JOIN productos p  ON dv.producto_id  = p.id
            LEFT JOIN  categorias c ON p.categoria_id  = c.id
            INNER JOIN ventas v     ON dv.venta_id     = v.id
            WHERE DATE(v.fecha) BETWEEN %s AND %s
            GROUP BY p.id, p.nombre, c.nombre
            ORDER BY SUM(dv.cantidad) DESC
            LIMIT 10
        """, [desde, hasta])
        top = cursor.fetchall()

        cursor.execute("""
            SELECT COALESCE(SUM(total),0), COUNT(*)
            FROM ventas
            WHERE DATE(fecha) BETWEEN %s AND %s
        """, [desde, hasta])
        resumen = cursor.fetchone()

    # Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{desde}_{hasta}.pdf"'

    doc    = SimpleDocTemplate(response, pagesize=letter,
                               topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story  = []

    # Título
    titulo = Paragraph(
        f'<b>Reporte de Ventas</b><br/>'
        f'<font size=11>Período: {desde} al {hasta}</font>',
        styles['Title']
    )
    story.append(titulo)
    story.append(Spacer(1, 0.25*inch))

    # Resumen general
    total_ingresos, total_ventas = resumen
    promedio = float(total_ingresos) / total_ventas if total_ventas > 0 else 0

    resumen_data = [
        ['Métrica',           'Valor'],
        ['Total ingresos',    f'${float(total_ingresos):,.2f}'],
        ['Total ventas',      str(total_ventas)],
        ['Promedio por venta',f'${promedio:,.2f}'],
    ]
    t_resumen = Table(resumen_data, colWidths=[3*inch, 2*inch])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#1e2a3a')),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 10),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f0f4ff'), colors.white]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',  (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    story.append(t_resumen)
    story.append(Spacer(1, 0.3*inch))

    # Top productos
    story.append(Paragraph('<b>Productos más vendidos</b>', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))

    top_data = [['Producto', 'Categoría', 'Unidades', 'Ingresos']]
    for row in top:
        top_data.append([
            row[0], row[1] or '—',
            str(row[2]), f'${float(row[3]):,.2f}'
        ])

    t_top = Table(top_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.5*inch])
    t_top.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#7b9cff')),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 9),
        ('ALIGN',       (2,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8faff'), colors.white]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',  (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0),(-1,-1), 5),
    ]))
    story.append(t_top)
    story.append(Spacer(1, 0.3*inch))

    # Detalle de ventas
    story.append(Paragraph('<b>Detalle de ventas</b>', styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))

    ventas_data = [['#', 'Empleado', 'Total', 'Fecha', 'Método']]
    for v in ventas:
        ventas_data.append([
            str(v[0]), v[1], f'${float(v[2]):,.2f}',
            str(v[3])[:16], v[4]
        ])

    t_ventas = Table(ventas_data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 1.8*inch, 1*inch])
    t_ventas.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#1e2a3a')),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 8),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f0f4ff'), colors.white]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',  (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))
    story.append(t_ventas)

    # Pie de página
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(
        f'<font size=8 color="#718096">Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}</font>',
        styles['Normal']
    ))

    doc.build(story)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_excel(request):
    desde = request.query_params.get('desde', '')
    hasta = request.query_params.get('hasta', '')

    # Obtener datos
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT v.id, u.nombre, v.total, v.fecha, v.metodo_pago
            FROM ventas v
            INNER JOIN usuarios u ON v.usuario_id = u.id
            WHERE DATE(v.fecha) BETWEEN %s AND %s
            ORDER BY v.fecha DESC
        """, [desde, hasta])
        ventas = cursor.fetchall()

        cursor.execute("""
            SELECT p.nombre, c.nombre, SUM(dv.cantidad), SUM(dv.subtotal)
            FROM detalle_ventas dv
            INNER JOIN productos p  ON dv.producto_id  = p.id
            LEFT JOIN  categorias c ON p.categoria_id  = c.id
            INNER JOIN ventas v     ON dv.venta_id     = v.id
            WHERE DATE(v.fecha) BETWEEN %s AND %s
            GROUP BY p.id, p.nombre, c.nombre
            ORDER BY SUM(dv.cantidad) DESC
        """, [desde, hasta])
        top = cursor.fetchall()

    # Crear Excel
    wb = openpyxl.Workbook()

    # ── Hoja 1: Ventas ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Ventas'

    # Estilo encabezado
    header_fill = PatternFill('solid', fgColor='1e2a3a')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    center      = Alignment(horizontal='center', vertical='center')

    ws1.append(['Reporte de Ventas', '', '', '', ''])
    ws1.merge_cells('A1:E1')
    ws1['A1'].font      = Font(bold=True, size=14, color='1e2a3a')
    ws1['A1'].alignment = center
    ws1.append([f'Período: {desde} al {hasta}', '', '', '', ''])
    ws1.merge_cells('A2:E2')
    ws1['A2'].alignment = center
    ws1.append([])

    encabezados = ['#', 'Empleado', 'Total ($)', 'Fecha', 'Método de pago']
    ws1.append(encabezados)
    for col, _ in enumerate(encabezados, 1):
        cell            = ws1.cell(row=4, column=col)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = center

    # Filas de datos
    alt_fill = PatternFill('solid', fgColor='f0f4ff')
    for i, v in enumerate(ventas):
        ws1.append([v[0], v[1], float(v[2]), str(v[3])[:16], v[4]])
        if i % 2 == 0:
            for col in range(1, 6):
                ws1.cell(row=5+i, column=col).fill = alt_fill

    # Anchos de columna
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 18

    # ── Hoja 2: Top Productos ───────────────────────────────
    ws2 = wb.create_sheet('Top Productos')
    ws2.append(['Productos más vendidos', '', '', ''])
    ws2.merge_cells('A1:D1')
    ws2['A1'].font      = Font(bold=True, size=14, color='1e2a3a')
    ws2['A1'].alignment = center
    ws2.append([])

    enc2 = ['Producto', 'Categoría', 'Unidades vendidas', 'Ingresos ($)']
    ws2.append(enc2)
    header_fill2 = PatternFill('solid', fgColor='7b9cff')
    for col, _ in enumerate(enc2, 1):
        cell           = ws2.cell(row=3, column=col)
        cell.fill      = header_fill2
        cell.font      = header_font
        cell.alignment = center

    for i, p in enumerate(top):
        ws2.append([p[0], p[1] or '—', int(p[2]), float(p[3])])
        if i % 2 == 0:
            for col in range(1, 5):
                ws2.cell(row=4+i, column=col).fill = alt_fill

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{desde}_{hasta}.xlsx"'
    wb.save(response)
    return response